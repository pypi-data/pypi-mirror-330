# Commented out section is for using OAuth instead of creating a service account

import os.path

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from datetime import datetime
# import mimetypes
import os, json

from genesis_bots.google_sheets.format_g_sheets import format_genesis_g_sheets

## test
from concurrent.futures import ThreadPoolExecutor
import time
from tenacity import retry, stop_after_attempt, wait_exponential

import openpyxl
import requests
from io import BytesIO

from genesis_bots.core.logging_config import logger
import re


# If modifying these scopes, delete the file token.json.
SCOPES = [
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/documents",
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets"
]

def column_to_number(letter: str) -> int:
    num = 0
    for char in letter:
        num = num * 26 + (ord(char.upper()) - ord('A') + 1)
    return num

def number_to_column(num: int) -> str:
    result = ""
    while num > 0:
        num -= 1
        result = chr(num % 26 + 65) + result
        num //= 26
    return result

def parse_cell_range(cell_range):
    cell_range = re.sub(r'[^A-Za-z0-9:]', '', cell_range)
    if cell_range.count(':') > 1:
        parts = cell_range.split(':')
        cell_range = parts[0] + ':' + parts[1]
    match = re.match(r"([A-Za-z]+)(\d+)(?::([A-Za-z]+)(\d+))?", cell_range)
    if not match:
        raise ValueError("Invalid cell range format")

    start_col, start_row, end_col, end_row = match.groups()
    start_col_num = column_to_number(start_col)
    start_row_num = int(start_row)

    if end_col and end_row:
        end_col_num = column_to_number(end_col)
        end_row_num = int(end_row)
    else:
        end_col_num = start_col_num
        end_row_num = start_row_num

    num_cells = (end_col_num - start_col_num + 1) * (end_row_num - start_row_num + 1)
    return start_col_num, start_row_num, end_col_num, end_row_num, num_cells


def insert_into_g_drive_file_version_table(self, data):
    """
    Insert a new row into the G_DRIVE_FILE_VERSION table.

    Args:
        data (dict): A dictionary containing the following keys:
            - g_file_id
            - g_file_name
            - g_file_type
            - g_file_parent_id
            - g_file_size
            - g_file_version
    """
    # Assuming you have a database connection established
    connection = self.connection
    cursor = connection.cursor()

    insert_query = f"""
    INSERT INTO {self.schema}.G_DRIVE_FILE_VERSION (g_file_id, g_file_name, g_file_type, g_file_parent_id, g_file_size, g_file_version)
    VALUES (%s, %s, %s, %s, %s, %s)
    """
    cursor.execute(
        insert_query,
        (
            data["g_file_id"],
            data["g_file_name"],
            data["g_file_type"],
            data["g_file_parent_id"],
            data["g_file_size"],
            data["g_file_version"],
        ),
    )

    connection.commit()
    cursor.close()
    return {"Success": True, "Message": "File version inserted."}


def update_g_drive_file_version_table(
    self, g_file_id, g_file_version, g_file_name, g_file_size, g_folder_id, g_file_type
):
    """
    Update the version of a file in the G_DRIVE_FILE_VERSION table.

    Args:
        g_file_id (str): The ID of the file.
        g_file_version (str): The new version of the file.
        g_file_name (str): The name of the file.
        g_file_size (str): The size of the file.
        g_folder_id (str): The ID of the folder containing the file.
        g_file_type (str): The type of the file.
    """
    connection = self.db_adapter.connection
    cursor = connection.cursor()

    # Check if the file ID exists in the table
    select_query = f"""
    SELECT COUNT(*)
    FROM {self.db_adapter.schema}.G_DRIVE_FILE_VERSION
    WHERE g_file_id = %s
    """
    cursor.execute(select_query, (g_file_id,))
    file_exists = cursor.fetchone()[0]

    if file_exists == 0:
        insert_query = f"""
        INSERT INTO {self.db_adapter.schema}.G_DRIVE_FILE_VERSION (g_file_id, g_file_version, g_file_name, g_file_size, g_file_parent_id, g_file_type)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(insert_query, (g_file_id, g_file_version, g_file_name, g_file_size, g_folder_id, g_file_type))
    else:
        update_query = f"""
        UPDATE {self.db_adapter.schema}.G_DRIVE_FILE_VERSION
        SET g_file_version = %s
        WHERE g_file_id = %s
        """
        cursor.execute(update_query, (g_file_version, g_file_id))

    connection.commit()
    cursor.close()

    return {"Success": True, "Message": "File version updated."}


def get_g_file_comments(user, file_id):
    """
    Get comments on a Google Sheets document.

    Args:
        file_id (str): The ID of the file.

    Returns:
        list: A list of comments on the document.
    """
    SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"

    try:
        # Authenticate using the service account JSON file
        creds = Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )

        service = build("drive", "v3", credentials=creds)

        # Get the comments on the document
        comments = (
            service.comments()
            .list(fileId=file_id, fields="comments(anchor, id,content,author(kind, displayName,emailAddress),replies(id,content,author(displayName,emailAddress),htmlContent))")
            .execute()
        )

        # Get the web link to the file
        file_metadata = service.files().get(fileId=file_id, fields="webViewLink").execute()
        file_url = file_metadata.get("webViewLink")

        flat_comments = []

        # Add the URL to each comment
        for comment in comments.get("comments", []):
            comment["url"] = f"{file_url}?comment={comment['id']}"
            for reply in comment.get("replies", []):
                reply["url"] = f"{file_url}?comment={comment['id']}&reply={reply['id']}"

            flat_comments.append(comment['content'])

        request = service.files().export_media(fileId=file_id, mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        fh = BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            print("Download %d%%" % int(status.progress() * 100))
        fh.seek(0)
        workbook = openpyxl.load_workbook(filename=fh, data_only=False)
        worksheet = workbook['Sheet1']
        res = []
        for i, row in enumerate(worksheet.iter_rows()):
            for j, cell in enumerate(row):
                if cell.comment:
                    try:
                        comment_index = flat_comments.index(cell.comment.text.split("\n", 1)[0])
                    except ValueError:
                        continue
                    comments['comments'][comment_index]["cellRC"] = number_to_column(j + 1).strip() + str(i + 1).strip()
                    comments["comments"][comment_index]["columnIndex"] = (
                        number_to_column(j + 1)
                    )

        return comments.get("comments", [])

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None


def add_reply_to_g_file_comment(
    file_id=None, comment_id=None, reply_content=None, g_file_comment_id=None, creds=None, user=None
):
    """
    Add a reply to a comment on a Google Drive file.

    Args:
        user (str): The user associated with the service account.
        file_id (str): The ID of the file.
        comment_id (str): The ID of the comment.
        reply_content (str): The content of the reply.

    Returns:
        dict: The created reply.
    """
    # if not file_id or not comment_id or not reply_content or not g_file_comment_id or (not creds and not user):
    #     raise Exception(
    #         "Missing credentials, user name, file ID, comment ID, or reply content."
    #     )

    SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"

    try:
        if not creds:
            # Authenticate using the service account JSON file
            creds = Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES
            )

        service = build("drive", "v3", credentials=creds)

        # Create the reply
        reply_body = {"content": reply_content}
        created_reply = (
            service.replies()
            .create(
                fileId=file_id,
                commentId=g_file_comment_id,
                body=reply_body,
                fields="id,content",
            )
            .execute()
        )

        print(f"Reply added: {created_reply['content']}")
        return created_reply

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None

def get_g_file_web_link(file_id, creds=None, user=None):
    """
    Get the web link to a file in Google Drive.

    Args:
        file_id (str): The ID of the file.

    Returns:
        str: The web link to the file.
    """
    # if not file_id or (not creds and not user):
    #     raise Exception("Missing credentials, user name, or file ID.")

    if not creds:
        SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
        try:
            # Authenticate using the service account JSON file
            creds = Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES
            )
        except Exception as e:
            print(f"Error loading credentials: {e}")
            return None

    try:
        service = build("drive", "v3", credentials=creds)

        # Get the file metadata including the webViewLink
        file_metadata = service.files().get(fileId=file_id, fields="name, webViewLink, parents").execute()

        return {
            "Success": True,
            "Name": file_metadata.get("name"),
            "URL": file_metadata.get("webViewLink"),
            "Folder ID": (
                file_metadata.get("parents")[0]
                if file_metadata.get("parents")
                else None
            ),
        }

    except Exception as e:
        return {"Success": False, "Error": str(e)}

def find_g_file_by_name(file_name, creds=None, user=None):
    """
    Find all files in Google Drive by their name.

    Args:
        file_name (str): The name of the file.

    Returns:
        dict: A list of file metadata if found, otherwise None.
    """
    # if not file_name or (not creds and not user):
    #     raise Exception("Missing credentials, user name, or file name.")

    if not creds:
        SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
        try:
            # Authenticate using the service account JSON file
            creds = Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES
            )
        except Exception as e:
            print(f"Error loading credentials: {e}")
            return None

    try:
        service = build("drive", "v3", credentials=creds)

        # Search for the files by name
        query = f"name='{file_name}'"
        response = service.files().list(q=query, fields="files(id, name, webViewLink, createdTime)").execute()
        files = response.get("files", [])

        if files:
            return {"Success": True, "Files": files}
        else:
            return {"Success": False, "Error": "Files not found"}

    except Exception as e:
        return {"Success": False, "Error": str(e)}

def get_g_folder_directory(folder_id, creds=None, user=None):
    """
    Get all files in a Google Drive folder.

    Args:
        folder_id (str): The ID of the folder.

    Returns:
        list: A list of files in the folder.
    """
    # if not folder_id or (not creds and not user):
    #     raise Exception("Missing credentials, user name, or folder ID.")
    logger.info(f"Entering get_g_folder_directory with folder_id: {folder_id}")

    if not creds:
        SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
        if not os.path.exists(SERVICE_ACCOUNT_FILE):
            logger.info(f"Service account file not found: {SERVICE_ACCOUNT_FILE}")
        try:
            # Authenticate using the service account JSON file
            creds = Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES
            )
            logger.info(f"Credentials loaded: {creds}")
        except Exception as e:
            logger.error(f"Error loading credentials: {e}")
            return False

    try:
        service = build("drive", "v3", credentials=creds)

        # Get the list of files in the folder
        query = f"'{folder_id}' in parents"
        response = service.files().list(q=query, fields="files(id, name)").execute()
        files = response.get("files", [])

        return {"Success": True, "File Names": files}

    except Exception as e:
        return {"Success": False, "Error": str(e)}

def add_g_file_comment(
    file_id=None,
    content=None,
    creds=None,
    user=None
):
    """
    Add a comment to a Google Drive file.

    Args:
        user (str): The user associated with the service account.
        file_id (str): The ID of the file.
        content (str): The content of the comment.

    Returns:
        dict: The created comment.
    """
    # if not file_id or not content or (not creds and not user):
    #     raise Exception(
    #         "Missing credentials, user name, file ID, or value."
        # )
    SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"

    try:
        if not creds:
            # Authenticate using the service account JSON file
            creds = Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES
            )
        service = build("drive", "v3", credentials=creds)

        # Create the comment
        body = {"content": content}
        created_comment = (
            service.comments()
            .create(fileId=file_id, body=body, fields="id,content")
            .execute()
        )

        print(f"Comment added: {created_comment['content']}")
        return created_comment

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None


def get_g_folder_web_link(folder_id, creds):
    """
    Get the web link to a folder in Google Drive.

    Args:
        folder_id (str): The ID of the folder.

    Returns:
        str: The web link to the folder.
    """
    try:
        # Authenticate using the service account JSON file
        service = build("drive", "v3", credentials=creds)

        # Get the folder metadata including the webViewLink
        folder = service.files().get(fileId=folder_id, fields="id, name, webViewLink").execute()

        # Print the folder details
        print(f"Folder ID: {folder.get('id')}")
        print(f"Folder Name: {folder.get('name')}")
        print(f"Web View Link: {folder.get('webViewLink')}")

        return folder.get("webViewLink")

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None


def get_g_file_version(g_file_id = None, creds = None, self = None):
    """
    Get the version number of a file in Google Drive.

    Args:
        file_id (str): The ID of the file.

    Returns:
        int: The version number of the file.
    """
    # if not g_file_id or (not self and not creds):
    #     raise Exception("Missing parameters in get_g_file_version - file id or user")

    if not creds:
        SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
        try:
            # Authenticate using the service account JSON file
            creds = Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES
            )
        except Exception as e:
            print(f"Error loading credentials: {e}")
            return None

    service = build("drive", "v3", credentials=creds)

    # Get the file metadata including the version number
    file_metadata = service.files().get(fileId=g_file_id, fields="version, name, size, parents").execute()

    version = file_metadata.get("version")
    file_name = file_metadata.get("name")
    file_size = file_metadata.get("size")
    parent_folder_id = file_metadata.get("parents")[0] if file_metadata.get("parents") else None
    g_file_type = 'sheet'

    update_g_drive_file_version_table(self, g_file_id, version, file_name, file_size, parent_folder_id, g_file_type)

    # Print the file version
    return version


# def upload_file_to_folder(path_to_file, parent_folder_id):
#     creds = Credentials.from_service_account_file(
#             SERVICE_ACCOUNT_FILE, scopes=SCOPES
#         )
#     service = build("drive", "v3", credentials=creds)

#     file_path = os.path(path_to_file)
#     filename = os.path.basename(file_path)
#     mime_type = mimetypes.guess_type(file_path)

#     file_metadata = {"name": filename}
#     if parent_folder_id:
#         file_metadata["parents"] = [parent_folder_id]

#     media = MediaFileUpload(file_path, mimetype=mime_type[0])
#     file = (
#         service.files().create(body=file_metadata, media_body=media, fields="id").execute()
#     )
#     print(f'File ID: "{file.get("id")}".')
#     return file.get("id")


def process_row(args):
    self, row, stage_column_index, stage_column_folder_ids, creds = args
    row_values = list(row.values())

    for j, row_value in enumerate(row_values):
        if isinstance(row_value, datetime):
            row_values[j] = row_value.strftime("%Y-%m-%d %H:%M:%S")
        elif len(stage_column_index) > 0 and j in stage_column_index and row_value:
            if len(row_value) < 1 or not row_value.startswith('@'):
                continue

            parts = row_value.split(".")
            path = parts[2].split("/")
            stage = path[0]

            file_contents = self.read_file_from_stage(
                parts[0].replace('@',''),
                parts[1],
                stage,
                "/".join(path[1:]) + '.' + parts[-1],
                True,
            )

            filename = path[-1] + '.' + parts[-1]
            stage_folder_id = stage_column_folder_ids[stage_column_index.index(j)]

            webLink = save_text_to_google_file_with_retry(
                self, stage_folder_id, filename, file_contents, creds
            )

            # Remove any quotes around the URL and filename to prevent formula errors
            webLink = webLink.replace('"', '') if webLink else ''
            filename = filename.replace('"', '')
            row_values[j] = f'=HYPERLINK("{webLink}")'

    return row_values

@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=4, max=10),
    retry_error_callback=lambda retry_state: None
)
def save_text_to_google_file_with_retry(*args, **kwargs):
    return save_text_to_google_file(*args, **kwargs)

def save_text_to_google_file(
    self, shared_folder_id, file_name, text = "No text in file", creds=None
):
    if not text or isinstance(text, dict):
        text = "No text received in save_text_to_google_file."

    if not creds:
        SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
        try:
            # Authenticate using the service account JSON file
            creds = Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES
            )
        except Exception as e:
            print(f"Error loading credentials: {e}")
            return None

    docs_service = build("docs", "v1", credentials=creds)
    drive_service = build("drive", "v3", credentials=creds)

    # Check if a document with the same name already exists in the shared folder
    query = f"'{shared_folder_id}' in parents and name='{file_name}' and mimeType='application/vnd.google-apps.document'"
    response = (
        drive_service.files().list(q=query, fields="files(id, name)").execute()
    )
    files = response.get("files", [])

    if files:
        for file in files:
            print(
                f"Deleting existing file: {file.get('name')} (ID: {file.get('id')})"
            )
            docs_service.files().delete(fileId=file.get("id")).execute()

    # Create a new document
    if not file_name:
        file_name = "genesis_" + datetime.now().strftime("%m%d%Y_%H:%M:%S")

    body = {"title": file_name}
    doc = docs_service.documents().create(body=body).execute()
    print("Created document with title: {0}".format(doc.get("title")))
    doc_id = doc.get("documentId")
    print(f"Document ID: {doc_id}")

    # Move the document to shared folder
    if shared_folder_id:
        file = (
            drive_service.files()
            .update(
                fileId=doc_id,
                addParents=shared_folder_id,
                fields="id, parents",
            )
            .execute()
        )
        print(f"File moved to folder: {file} | Parent folder {file['parents'][0]}")

    # Verify the new document exists in Google Drive
    try:
        file_verify = (
            drive_service.files()
            .get(fileId=doc_id, fields="id, name, parents, webViewLink")
            .execute()
        )
        print(f"File store confirmed: {file_verify}")
    except:
        raise Exception("Error creating document in Google Drive")

    parent = (
        drive_service.files().get(fileId=shared_folder_id, fields="id, name").execute()
    )
    print(f"Parent folder name: {parent.get('name')} (ID: {parent.get('id')})")

    requests = [{"insertText": {"location": {"index": 1}, "text": text}}]

    result = (
        docs_service.documents()
        .batchUpdate(documentId=doc_id, body={"requests": requests})
        .execute()
    )

    print("Document content updated: ", result)

    # Add to G_DRIVE_FILE_VERSION table
    g_file_version_data = {
        "g_file_id": doc_id,
        "g_file_name": file_name,
        "g_file_type": "application/vnd.google-apps.document",
        "g_file_parent_id": shared_folder_id,
        "g_file_size": str(len(text)),
        "g_file_version": "1"
    }

    insert_into_g_drive_file_version_table(self, g_file_version_data)

    return {
            "Success": True,
            "URL": file_verify.get("webViewLink"),
        }


def create_folder_in_folder(folder_name, parent_folder_id, user):
    SERVICE_ACCOUNT_FILE = f'g-workspace-credentials.json'
    creds = Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )
    service = build("drive", "v3", credentials=creds)

    file_metadata = {
        "name": folder_name,
        "parents": [parent_folder_id],
        "mimeType": "application/vnd.google-apps.folder",
    }

    file = service.files().create(body=file_metadata, fields="id").execute()

    print(f'Folder ID: {file.get("id")} | Folder name: {folder_name}')

    return file.get("id")


def export_to_google_docs(text: str = 'No text received.', shared_folder_id: str = None, user =None, file_name = None):
    """
    Creates new file in Google Docs named Genesis_mmddyyy_hh:mm:ss from text string
    """
    pass
    # if not user:
    #     raise Exception("User not specified for google drive conventions.")

    # SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
    # try:
    #     # Authenticate using the service account JSON file
    #     creds = Credentials.from_service_account_file(
    #         SERVICE_ACCOUNT_FILE, scopes=SCOPES
    #     )
    #     docs_service = build("docs", "v1", credentials=creds)
    #     drive_service = build("drive", "v3", credentials=creds)

    #     # Check if a document with the same name already exists in the shared folder
    #     query = f"'{shared_folder_id}' in parents and name='{file_name}' and mimeType='application/vnd.google-apps.document'"
    #     response = drive_service.files().list(q=query, fields="files(id, name)").execute()
    #     files = response.get("files", [])

    #     if files:
    #         for file in files:
    #             print(f"Deleting existing file: {file.get('name')} (ID: {file.get('id')})")
    #             drive_service.files().delete(fileId=file.get("id")).execute()

    #     # Create a new document
    #     if not file_name:
    #         file_name = "genesis_" + datetime.now().strftime("%m%d%Y_%H:%M:%S")

    #     body = {"title": file_name}
    #     doc = docs_service.documents().create(body=body).execute()
    #     print("Created document with title: {0}".format(doc.get("title")))
    #     doc_id = doc.get("documentId")
    #     print(f"Document ID: {doc_id}")

    #     # Move the document to shared folder
    #     if shared_folder_id:
    #         file = (
    #             drive_service.files()
    #             .update(
    #                 fileId=doc_id,
    #                 addParents=shared_folder_id,
    #                 fields="id, parents",
    #             )
    #             .execute()
    #         )
    #         print(f"File moved to folder: {file} | Parent folder {file['parents'][0]}")

    #     # Verify the new document exists in Google Drive
    #     try:
    #         file_verify = (
    #             drive_service.files()
    #             .get(fileId=doc_id, fields="id, name, parents, webViewLink")
    #             .execute()
    #         )
    #         print(f"File store confirmed: {file_verify}")
    #     except:
    #         raise Exception("Error creating document in Google Drive")

    #     parent = (
    #         drive_service.files().get(fileId=shared_folder_id, fields="id, name").execute()
    #     )
    #     print(f"Parent folder name: {parent.get('name')} (ID: {parent.get('id')})")

    #     if not text:
    #         text = 'No text received from Snowflake stage.'

    #     requests = [{"insertText": {"location": {"index": 1}, "text": text}}]

    #     result = (
    #         docs_service.documents()
    #         .batchUpdate(documentId=doc_id, body={"requests": requests})
    #         .execute()
    #     )

    #     print("Document content updated: ", result)

    #     return file_verify.get("webViewLink")

    # except HttpError as err:
    #     print(err)
    #     return None


# def create_google_sheet_no_v4(self, shared_folder_id, title, data):
#     """
#     Creates a Google Sheet with the given title and table data and moves it
#     from the service account to the shared folder.
#     Loads pre-authorized user credentials from the environment.
#     """
#     # if not self.user:
#     #     raise Exception("User not specified for google drive conventions.")

#     SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
#     try:
#         # Authenticate using the service account JSON file
#         creds = Credentials.from_service_account_file(
#             SERVICE_ACCOUNT_FILE, scopes=SCOPES
#         )
#     except Exception as e:
#         print(f"Error loading credentials: {e}")
#         return None

#     try:
#         # service = build("sheets", "v4", credentials=creds)
#         service = build("drive", "v3", credentials=creds)

#         new_workbook = openpyxl.Workbook()
#         new_worksheet = new_workbook.active

#         temp_file_path = "temp_google_sheet.xlsx"
#         new_workbook.save(temp_file_path)

#         i = 0
#         for id, obj in enumerate(data):
#             j = 0
#             for key, value in enumerate(obj):
#                 new_worksheet.cell(row=i, column=j, value=value)
#                 j += 1
#             i += 1

#         # Save the workbook to a temporary file
#         temp_file_path = "temp_google_sheet.xlsx"
#         new_workbook.save(temp_file_path)

#         # Upload the file back to Google Drive
#         # service = result['service'] #build("drive", "v3", credentials=creds)
#         media = MediaFileUpload(
#             temp_file_path,
#             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         )
#         file = service.files().update(fileId=spreadsheet_id, media_body=media).execute()

#         print(f"File ID: {file.get('id')}")
#         return {
#             "Success": True,
#             "updatedCells": result.get("updatedCells"),
#             "file_id": file.get("id"),
#         }

#     except Exception as e:
#         print(f"An error occurred: {str(e)}")
#         return {
#             "Success": False,
#             "Error": str(e),
#         }

def create_google_sheet_from_export(self, shared_folder_id, title, data):
    """
    Creates a Google Sheet with the given title and table data and moves it
    from the service account to the shared folder.
    Loads pre-authorized user credentials from the environment.
    """
    # if not self.user:
    #     raise Exception("User not specified for google drive conventions.")

    SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
    try:
        # Authenticate using the service account JSON file
        creds = Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )
    except Exception as e:
        print(f"Error loading credentials: {e}")
        return None

    try:
        # service = build("sheets", "v4", credentials=creds)
        drive_service = build("drive", "v3", credentials=creds)

        service = build("sheets", "v4", credentials=creds)

        spreadsheet = {"properties": {"title": title}}
        spreadsheet = (
            service.spreadsheets()
            .create(body=spreadsheet, fields="spreadsheetId")
            .execute()
        )

        ss_id = spreadsheet.get("spreadsheetId")
        print(f"Spreadsheet ID: {ss_id}")
        keys = list(data[0].keys())
        columns = [keys]

        # Check for stage links
        stage_column_index = [i for i, key in enumerate(keys) if key.endswith("_STAGE_LINK")]
        stage_column_folder_names = [key.replace("_STAGE_LINK", "") for key in keys if key.endswith("_STAGE_LINK")]
        stage_column_folder_ids = []

        # Create folder top level folder
        top_level_folder_id = create_folder_in_folder(
            title + "(" + datetime.now().strftime("%m%d%Y_%H:%M:%S") + ")",
            shared_folder_id,
            self.user
        )

        if len(stage_column_folder_names) > 0:
            # Create sub-folders
            for stage_column_folder in stage_column_folder_names:
                stage_column_folder_ids.append(
                    create_folder_in_folder(
                        stage_column_folder,
                        top_level_folder_id,
                        self.user
                    )
                )

        # Process rows in smaller batches
        batch_size = 10  # Adjust based on your needs
        max_workers = 5  # Reduced number of concurrent workers
        processed_rows = []

        for i in range(0, len(data), batch_size):
            batch = data[i:i + batch_size]
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                row_args = [(self, row, stage_column_index, stage_column_folder_ids, creds)
                    for row in batch]
                batch_results = list(executor.map(process_row, row_args))
                processed_rows.extend(batch_results)
                time.sleep(1)  # Add delay between batches

        # Add header and processed rows to columns
        columns = [keys] + processed_rows

        spreadsheet = {"properties": {"title": title}}
        spreadsheet = (
            service.spreadsheets()
            .create(body=spreadsheet, fields="spreadsheetId")
            .execute()
        )

        ss_id = spreadsheet.get("spreadsheetId")

        width_10 = chr(65 + len(columns[0]) % 26)
        width_1 = chr(64 + len(columns[0]) // 26) if len(columns[0]) > 25 else ''
        width = width_10 + width_1
        cell_range = f"Sheet1!A1:{width}{len(columns)}"
        print(f"\n\nRange name: {cell_range} | {len(columns[0])} | {len(columns)}\n\n")
        body = {
                "values": columns
                }

        result = (
            service.spreadsheets()
            .values()
            .update(
                spreadsheetId=ss_id,
                range=cell_range,
                valueInputOption='USER_ENTERED',
                body=body,
            )
            .execute()
        )
        print(f"{result.get('updatedCells')} cells created.")

        # Apply formatting
        service.spreadsheets().batchUpdate(
            spreadsheetId=ss_id,
            body={"requests": format_genesis_g_sheets(columns)}
        ).execute()

        # Move the document to shared folder
        if top_level_folder_id:
            file = (
                drive_service.files()
                .update(
                    fileId=ss_id,
                    addParents=top_level_folder_id,
                    fields="id, webViewLink, parents",
                )
                .execute()
            )
            print(f"File moved to folder - File ID: {file['id']} | Folder ID {file['parents'][0]}")

        # Test only - read file contents to confirm write
        # results = read_g_sheet(ss_id, cell_range, creds)
        # print(f"Results from storing, then reading sheet: {results}")

        folder_url = get_g_folder_web_link(top_level_folder_id, creds)
        file_url = file.get("webViewLink")

        g_file_version_data = {
            "g_file_id": ss_id,
            "g_file_name": title,
            "g_file_type": "sheet",
            "g_file_parent_id": top_level_folder_id,
            "g_file_size": None,
            "g_file_version": "1",
        }

        insert_into_g_drive_file_version_table(self, g_file_version_data)

        return {"Success": True, "file_id": spreadsheet.get("spreadsheetId"), "file_url": file_url, "folder_url": folder_url}

    except HttpError as error:
        print(f"An error occurred: {error}")
        return error

def create_g_sheet_v4(g_sheet_values, g_sheet_name = "Google Sheet", creds=None, user=None):
    """
    Create a Google Sheet with the given values.
    Load pre-authorized user credentials from the environment.
    """
    if not user:
        raise Exception("User not specified for google drive conventions.")

    SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
    try:
        # Authenticate using the service account JSON file
        creds = Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )
    except Exception as e:
        print(f"Error loading credentials: {e}")
        return None

    try:
        service = build("sheets", "v4", credentials=creds)

        # Create the Google Sheet
        spreadsheet = {"properties": {"title": g_sheet_name}}
        spreadsheet = (
            service.spreadsheets()
            .create(body=spreadsheet, fields="spreadsheetId")
            .execute()
        )

        ss_id = spreadsheet.get("spreadsheetId")
        print(f"Spreadsheet ID: {ss_id}")

        # Prepare the body for the update request
        body = {
            "values": g_sheet_values
        }

        # Update the Google Sheet with the new values
        result = (
            service.spreadsheets()
            .values()
            .update(
                spreadsheetId=ss_id,
                range="Sheet1!A1",
                valueInputOption="USER_ENTERED",
                body=body,
            )
            .execute()
        )

        print(f"{result.get('updatedCells')} cells created.")

        return {
            "Success": True,
            "file_id": spreadsheet.get("spreadsheetId"),
        }

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return {
            "Success": False,
            "Error": str(e),
        }

def write_g_sheet_cell_v3(spreadsheet_id=None, cell_range=None, value=None, creds=None, user=None):
    # if not spreadsheet_id or not cell_range or (not creds and not user):
    #     raise Exception(
    #         "Missing credentials, user name, spreadsheet ID, or cell_range name."
    #     )
    logger.info(f"Entering write_g_sheet with ss_id: {spreadsheet_id}")

    if not creds:
        SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
        try:
            # Authenticate using the service account JSON file
            logger.info(f"Try auth: {spreadsheet_id}")
            creds = Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES
            )
            logger.info(f"Auth success: {spreadsheet_id}")
        except Exception as e:
            logger.info(f"Error loading credentials: {spreadsheet_id}")
            print(f"Error loading credentials: {e}")
            return None

    service = build("drive", "v3", credentials=creds)

    result = read_g_sheet(spreadsheet_id, cell_range, creds, user)

    start_col, start_row, end_col, end_row, num_cells = (
        parse_cell_range(cell_range)
    )

    value_arr = value.split(",")

    if len(value_arr) != num_cells:
        raise ValueError("Number of values does not match the number of cells in cell_range")

    print(f"Start Column: {start_col}, Start Row: {start_row}, End Column: {end_col}, End Row: {end_row}")

    # Update the result['cell_values'] with the values from value_arr
    index = 0
    for col in range(start_col - 1, end_col):
        for row in range(start_row - 1, end_row):
            result['cell_values'][row][col] = value_arr[index]
            index += 1

    # # Prepare the body for the update request
    # body = {
    #     "values": result['cell_values'][start_row - 1:end_row]
    # }

    # Update the Google Sheet with the new values - DOES NOT WORK ON SPCS DOCKER
    # service = build("sheets", "v4", credentials=creds)
    # result = (
    #     service.spreadsheets()
    #     .values()
    #     .update(
    #         spreadsheetId=spreadsheet_id,
    #         range=cell_range,
    #         valueInputOption='USER_ENTERED',
    #         body=body,
    #     )
    #     .execute()
    # )
    # Write the updated values back to the Google Sheet using openpyxl
    try:
        # Create a new workbook and worksheet
        new_workbook = openpyxl.Workbook()
        new_worksheet = new_workbook.active

        # Write the updated cell values to the new worksheet
        for row_idx, row in enumerate(result['cell_values'], start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                new_worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

        # Save the workbook to a temporary file
        temp_file_path = "temp_google_sheet.xlsx"
        new_workbook.save(temp_file_path)

        # Upload the file back to Google Drive
        # service = result['service'] #build("drive", "v3", credentials=creds)
        media = MediaFileUpload(temp_file_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        file = service.files().update(fileId=spreadsheet_id, media_body=media).execute()

        print(f"File ID: {file.get('id')}")
        return {
            "Success": True,
            "updatedCells": result.get("updatedCells"),
            "file_id": file.get("id"),
        }

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return {
            "Success": False,
            "Error": str(e),
        }


def write_g_sheet_cell_v4(
    spreadsheet_id=None, cell_range=None, value=None, creds=None, user=None
):
    # if not spreadsheet_id or not cell_range or (not creds and not user):
    #     raise Exception(
    #         "Missing credentials, user name, spreadsheet ID, or cell_range name."
    #     )
    if not creds:
        SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
        try:
            # Authenticate using the service account JSON file
            creds = Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES
            )
        except Exception as e:
            print(f"Error loading credentials: {e}")
            return None

    service = build("sheets", "v4", credentials=creds)

    body = {"values": [[value]]}

    result = (
        service.spreadsheets()
        .values()
        .update(
            spreadsheetId=spreadsheet_id,
            range=cell_range,
            valueInputOption="USER_ENTERED",
            body=body,
        )
        .execute()
    )
    return {
        "Success": True,
        "updatedCells": result.get("updatedCells"),
    }


def read_g_sheet(spreadsheet_id=None, cell_range=None, creds=None, user=None):
    """
    Reads the content of a Google Sheet.
    Load pre-authorized user credentials from the environment.
    """
    logger.info(f"Entering read_g_sheet with ss_id: {spreadsheet_id}")

    if not creds:
        SERVICE_ACCOUNT_FILE = f"g-workspace-credentials.json"
        try:
            # Authenticate using the service account JSON file
            logger.info(f"Try auth: {spreadsheet_id}")
            creds = Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES
            )
            logger.info(f"Auth success: {spreadsheet_id}")
        except Exception as e:
            logger.info(f"Error loading credentials: {spreadsheet_id}")
            print(f"Error loading credentials: {e}")
            return None
    try:
        service = build("drive", "v3", credentials=creds)

        request = service.files().export_media(
            fileId=spreadsheet_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        fh = BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            print("Download %d%%" % int(status.progress() * 100))
        fh.seek(0)
        workbook = openpyxl.load_workbook(filename=fh, data_only=False)
        worksheet = workbook[workbook.sheetnames[0]]

        # Extract the content of the worksheet
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(list(row))

        return {
            "Success": True,
            "cell_values": rows,
            "service": service,
        }
    except Exception as error:
        print(f"An error occurred: {error}")
        logger.info(f"HTTPError in read sheet: {error} - {spreadsheet_id}")
        return error
