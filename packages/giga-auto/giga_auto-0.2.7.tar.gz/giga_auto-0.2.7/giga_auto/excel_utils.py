

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def modified_excel_data(**data):
    """
    修改excel文件中的单个数据
    data = {
    "file": "文件路径",
    "sheet_name": "表名",
    "item_name": "A1",
    "item_value": "value"
    }
    """
    try:
        wb = load_workbook(data['file'])
        sheet = wb[data['sheet_name']]
        sheet[data['item_name']] = data['item_value']
        wb.save(data['file'])
    except InvalidFileException:
        raise InvalidFileException("无法打开文件：文件格式不正确或已损坏。")
    except FileNotFoundError:
        raise FileNotFoundError("文件未找到，请检查文件路径是否正确。")
    except Exception as e:
        raise Exception(f"修改excel文件遇到未知错误：{e}")


def modified_excel_datas(**data):
    """
    修改excel文件中的多个数据
    data = {
    "file": "文件路径",
    "sheet_name": "表名",
    "modified_data": {
        "A1": "value1",
        "B1": "value2",
        "C1": "value3"
    }
    }
    """
    try:
        wb = load_workbook(data['file'])
        sheet = wb[data['sheet_name']]
        modified_data = data['modified_data']
        for key,value in modified_data.items():
            sheet[key] = value
        wb.save(data['file'])
    except InvalidFileException:
        raise InvalidFileException("无法打开文件：文件格式不正确或已损坏。")
    except FileNotFoundError:
        raise FileNotFoundError("文件未找到，请检查文件路径是否正确。")
    except Exception as e:
        raise Exception(f"修改excel文件遇到未知错误：{e}")