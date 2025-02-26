from dataclasses import dataclass
from typing import Type

import openpyxl
import csv

from . import utils


class Base:
    def __init__(self, filename: str | None = None, mode: str = 'a'):
        self.filename = filename
        self.mode = mode

    def _new(self):
        pass

    def _open(self, filename: str):
        pass

    def _save(self, filename: str):
        pass

    def write_header(self, cls: Type[dataclass]):
        pass

    def write_record(self, cls: dataclass):
        pass

    def read(self) -> list[list[str]]:
        pass

    def __enter__(self):
        if self.mode in ('r', 'a') and self.filename:
            self._open(self.filename)
        else:
            self._new()
        return self

    def __exit__(self, type, val, traceback):
        self._save(self.filename)


class CSV(Base):
    def __init__(self,
                 filename: str | None = None,
                 mode: str = 'w',
                 delimiter: str = ';',
                 newline: str = '',
                 quotechar: str = '|'):
        
        super().__init__(filename, mode)
        self.delimiter = delimiter
        self.newline = newline
        self.quotechar = quotechar
        self.content: list = []

    def _new(self):
        self.content: list = []

    def _save(self, filename: str):
        with open(filename, 'w', newline=self.newline) as f:
            writer = csv.writer(f, delimiter=self.delimiter, quotechar=self.quotechar, quoting=csv.QUOTE_MINIMAL)
            writer.writerows(self.content)

    def _open(self, filename: str):
        try:
            with open(filename, newline=self.newline) as f:
                reader = csv.reader(f, delimiter=self.delimiter, quotechar=self.quotechar)
                for row in reader:
                    self.content.append(row)
        except FileNotFoundError:
            pass

    def write_header(self, cls: Type[dataclass]):
        self.content.append(tuple(utils.get_dataclass_fields(cls)))
    
    def write_record(self, cls: dataclass):
        fields = utils.get_dataclass_fields(cls)
        self.content.append([getattr(cls, i) for i in fields])

    def read(self) -> list[list[str]]:
        return self.content


class Excel(Base):
    def __init__(self, filename: str | None = None, mode: str = 'w'):
        super().__init__(filename, mode)
        self.wb: openpyxl.Workbook | None = None
        self.ws = None

    def _new(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def _save(self, filename: str):
        self.wb.save(filename)

    def _open(self, filename: str):
        self.wb: openpyxl.Workbook | None = openpyxl.load_workbook(filename)
        self.ws = self.wb.active

    def write_header(self, cls: Type[dataclass]):
        self.ws.append(tuple(utils.get_dataclass_fields(cls)))
    
    def write_record(self, cls: dataclass):
        fields = utils.get_dataclass_fields(cls)
        self.ws.append([getattr(cls, i) for i in fields])

    def read(self) -> list[list[str]]:
        res = []
        for row in self.ws.iter_rows():
            res.append([cell.value for cell in row])
        return res
