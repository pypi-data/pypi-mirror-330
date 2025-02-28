from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from pathlib import PureWindowsPath
from urllib.parse import unquote
import re


defined_ranges = []

def build_defined_ranges(defined_names):
    defined_names = [x for x in defined_names.items() if not x[0].startswith("_xlchart")]
    for name in defined_names:
        defn = name[0]
        for title, coord in name[1].destinations:
            defined_ranges.append({
                'name': defn,
                'sheet': title,
                'range': CellRange(coord),
            })

def is_defined(cell):
    sheet_name = cell.parent.title
    coord = cell.coordinate
    ranges = [x['name'] for x in defined_ranges if x['sheet'] == sheet_name and not x['range'].isdisjoint(CellRange(coord))]
    return ranges

def extract(formula, curr_sheet, links):
    ref_pattern = r"(?:(?: *'?\[([^\]]+)\])?([^'=,]+)'?\!)?([A-Z|$]+\d+(?::[A-Z|$]+\d+)?)"
    matches = re.findall(ref_pattern, formula)
    refs = []

    for match in matches:
        ref = {}
        workbook, sheet, cell = match
        if (workbook):
            ref["file"] = PureWindowsPath(unquote(links[int(workbook)-1].file_link.Target)).name
        if (sheet):
            ref["sheet"] = sheet
        else:
            ref["sheet"] = curr_sheet.title
        ref["cell"] = cell
        refs.append(ref)

    return refs

def getNames(cell):
    defd = is_defined(cell)
    if (len(defd) > 0):
        return tuple(defd)
    sheet = cell.parent
    row_label = None
    col_label = None
    i = cell.row
    j = cell.col_idx
    while (j > 1):
        if sheet.cell(i, j).data_type == 's' or sheet.cell(i, j).data_type == 'd':
            row_label = str(sheet.cell(i, j).value)
            break
        j -= 1
    j = cell.col_idx
    while (i > 1):
        if sheet.cell(i, j).data_type == 's' or sheet.cell(i, j).data_type == 'd':
            col_label = str(sheet.cell(i, j).value)
            break
        i -= 1
    return (row_label, col_label)

def parse_excel_formulas(sheet, links):
    formulas = {}

    for row in sheet.iter_rows():
        for cell in row:
            if (isinstance(cell.value, str) and cell.value.startswith('=')):
                names = getNames(cell)
                formulas[cell.coordinate] = {
                    "names": names,
                    "formula": cell.value,
                    "references": extract(cell.value, sheet, links),
                }

    return formulas

def parse_all_sheets(file_path):
    wb = load_workbook(file_path, data_only=False)
    build_defined_ranges(wb.defined_names)
    all_refs = {}

    for sheet in wb.sheetnames:
        all_refs[sheet] = {
            "items": parse_excel_formulas(wb.get_sheet_by_name(sheet), wb._external_links),
        }

    return all_refs

def parse_single_sheet(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=False)
    sheet = wb[sheet_name]
    build_defined_ranges(wb.defined_names)

    return parse_excel_formulas(sheet, wb._external_links)
