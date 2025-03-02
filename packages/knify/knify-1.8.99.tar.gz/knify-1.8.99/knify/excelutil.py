#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author: qicongsheng
from typing import Callable

import xlrd
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import listutil
from . import objutil


class Header:
    def __init__(self, index: int, name: str | None,
                 transformer: Callable[[object], object] = None):
        self.index = index
        self.name = name
        self.transformer = transformer


class HeaderBuilder:
    def __init__(self):
        self.default_transformer = None
        self.headers = []

    def set_default_transformer(self,
                                transformer: Callable[[object], object] = None) -> object:
        self.default_transformer = transformer
        return self

    def set_names(self, names: list[str] = None) -> object:
        for index_, name in enumerate(names):
            self.headers.append(Header(index_, name, self.default_transformer))
        return self

    def set_transformer(self, name: str,
                        transformer: Callable[[object], object] = None) -> object:
        for header in self.headers:
            if name == header.name:
                header.transformer = transformer
        return self

    def append(self, index: int, name: str | None,
               transformer: Callable[[object], object] = None) -> object:
        target_index = objutil.default_if_none(index, len(self.headers))
        target_transformer = objutil.default_if_none(transformer,
                                                     self.default_transformer)
        self.headers.append(Header(target_index, name, target_transformer))
        return self

    def to_headers(self) -> list[Header]:
        return self.headers


def read_excel(file_path: str, sheet: str | int | None = 0,
               headers: list[Header] | None = None, start_row: int = 1,
               header_row: int = 0) -> list[object]:
    results = []
    workbook = load_workbook(filename=file_path)
    sheet_ = workbook[sheet] if isinstance(sheet, str) else workbook[
        workbook.sheetnames[sheet]]
    headers_ = [cell.value for cell in sheet_[header_row + 1]]
    for row_idx, row in enumerate(sheet_.rows):
        if row_idx < start_row:
            continue
        result = {}
        for header_idx, header_ in enumerate(headers_):
            # 没有传入headers,使用默认header
            if listutil.is_empty(headers):
                result[header_] = row[header_idx].value
            # 传入了headers
            else:
                header = listutil.find_first(
                    list(filter(lambda h_: h_.index == header_idx, headers)))
                if header is None:
                    continue
                else:
                    col_name = objutil.default_if_none(header.name, header_)
                    cell_value = row[header_idx].value
                    cell_value = cell_value if header.transformer is None else header.transformer(
                        row[header_idx].value)
                    result[col_name] = cell_value
        if objutil.has_keys(result):
            results.append(result)
    return results


def read_headers(file_path: str, sheet: str | int | None = 0,
                 header_row: int = 0):
    workbook = load_workbook(filename=file_path)
    sheet_ = workbook[sheet] if isinstance(sheet, str) else workbook[
        workbook.sheetnames[sheet]]
    return [cell.value for cell in sheet_[header_row + 1]]


def load_excel_data(file_path):
    """通用Excel加载函数，支持xls和xlsx格式"""
    if file_path.endswith('.xls'):
        wb = xlrd.open_workbook(file_path)
        sheet = wb.sheet_by_index(0)
        headers = sheet.row_values(0)
        data = [sheet.row_values(i) for i in range(1, sheet.nrows)]
        return headers, data, None  # xls格式不获取列宽
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active
        headers = [cell.value for cell in next(sheet.iter_rows())]
        data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]

        # 获取列宽信息
        column_widths = []
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            column_dim = sheet.column_dimensions.get(col_letter)
            if column_dim and column_dim.width is not None:
                column_widths.append(column_dim.width)
            else:
                column_widths.append(8.43)  # Excel默认列宽
        return headers, data, column_widths


def compare(file1_path, file2_path, output_path, key_column,
            file1_alias="文件1", file2_alias="文件2"):
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00',
                              fill_type='solid')
    header_fill = PatternFill(start_color='AFEEEE', end_color='AFEEEE',
                              fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 加载数据和列宽
    headers1, data1_rows, col_widths1 = load_excel_data(file1_path)
    headers2, data2_rows, _ = load_excel_data(file2_path)

    if headers1 != headers2:
        raise ValueError("两个文件的表头不一致")
    if key_column not in headers1:
        raise ValueError(f"主键列 {key_column} 不存在")

    key_index = headers1.index(key_column)
    other_headers = [h for h in headers1 if h != key_column]

    # 构建数据字典
    def build_data_dict(data_rows):
        return {row[key_index]: row for row in data_rows}

    data1 = build_data_dict(data1_rows)
    data2 = build_data_dict(data2_rows)

    # 创建结果工作簿
    result_wb = Workbook()
    result_ws = result_wb.active
    result_ws.title = "对比结果"

    # ========== 构建表头 ==========
    # 第一行结构：主键列 + 合并列名
    header_row1 = [key_column] + [h for h in other_headers for _ in (0, 0)]
    result_ws.append(header_row1)

    # 第二行结构：空主键 + 文件标识
    header_row2 = [""] + [f for h in other_headers for f in
                          (file1_alias, file2_alias)]
    result_ws.append(header_row2)

    # 合并主键列单元格（纵向合并）
    result_ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    result_ws.cell(1, 1).alignment = Alignment(horizontal='center',
                                               vertical='center')

    # 合并其他列单元格（横向合并）
    col_pos = 2
    for h in other_headers:
        result_ws.merge_cells(
            start_row=1,
            start_column=col_pos,
            end_row=1,
            end_column=col_pos + 1
        )
        result_ws.cell(1, col_pos).value = h
        col_pos += 2

    # 设置表头样式
    for cell in result_ws[1]:
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    for cell in result_ws[2]:
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # ====== 冻结窗格和筛选 ======
    result_ws.freeze_panes = 'B3'  # 冻结第一列和前两行
    result_ws.auto_filter.ref = f"A2:{get_column_letter(result_ws.max_column)}2"  # 第二行添加筛选

    # ========== 数据对比处理 ==========
    all_keys = sorted(set(data1.keys()) | set(data2.keys()))

    for key in all_keys:
        row1 = data1.get(key)
        row2 = data2.get(key)

        # 构建结果行
        result_row = [key]
        for h in other_headers:
            idx = headers1.index(h)
            val1 = row1[idx] if row1 else None
            val2 = row2[idx] if row2 else None
            result_row.extend([val1, val2])

        result_ws.append(result_row)
        current_row = result_ws.max_row

        # 设置数据行边框
        for col in range(1, result_ws.max_column + 1):
            result_ws.cell(row=current_row, column=col).border = thin_border

        # 判断行存在情况
        exists1 = key in data1
        exists2 = key in data2

        # 标记差异逻辑
        if exists1 and exists2:
            # 逐列比较差异
            for col_idx, h in enumerate(other_headers, 1):
                idx = headers1.index(h)
                val1 = row1[idx]
                val2 = row2[idx]
                if val1 != val2:
                    target_col = 1 + col_idx * 2
                    result_ws.cell(current_row, target_col).fill = yellow_fill
                    # result_ws.cell(current_row, target_col+1).fill = yellow_fill
        else:
            # 整行标黄逻辑
            for col_idx, h in enumerate(other_headers, 1):
                target_col = 1 + col_idx * 2
                if exists1 and not exists2:  # 只存在文件1
                    result_ws.cell(current_row, target_col).fill = yellow_fill
                elif exists2 and not exists1:  # 只存在文件2
                    result_ws.cell(current_row, target_col).fill = yellow_fill

    # ========== 设置列宽 ==========
    if col_widths1:
        # 设置主键列宽
        key_width = col_widths1[key_index]
        result_ws.column_dimensions['A'].width = key_width

        # 设置其他列宽
        for idx, width in enumerate(col_widths1):
            if headers1[idx] == key_column:
                continue
            # 计算结果文件列位置
            pos_in_other = other_headers.index(headers1[idx])
            result_col = 2 + pos_in_other * 2
            # 设置两列宽度
            result_ws.column_dimensions[get_column_letter(result_col)].width = width
            result_ws.column_dimensions[
                get_column_letter(result_col + 1)].width = width

    result_wb.save(output_path)
